# -*- coding: utf-8 -*-
import re
import sys
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

# ========= KONFIG =========
DEFAULT_INPUT = "pracujacy_przeds.xlsx"
SHEET_OUT = "SUMY_CHRONO"
OUTPUT_SUFFIX = "_sumy_chrono"
OUTPUT_PERIOD_FORMAT = "MM.YYYY"   # "MM.YYYY" albo "YYYY-MM"
# ==========================

MONTH_MAP = {
    "styczeń": 1, "styczen": 1,
    "luty": 2,
    "marzec": 3,
    "kwiecień": 4, "kwiecien": 4,
    "maj": 5,
    "czerwiec": 6,
    "lipiec": 7,
    "sierpień": 8, "sierpien": 8,
    "wrzesień": 9, "wrzesien": 9,
    "październik": 10, "pazdziernik": 10,
    "listopad": 11,
    "grudzień": 12, "grudzien": 12,
}
YEAR_RE = re.compile(r"^(19|20)\d{2}$")


def to_number_pl(x):
    """Obsługuje: '523 947', '1 623 888', '4,9', '120000', NBSP, itp."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "brak", "-"}:
        return np.nan

    s = s.replace("\u00A0", " ").replace(" ", "")  # usuń spacje i NBSP
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)
    if s in {"", "-", ".", "-."}:
        return np.nan

    try:
        return float(s)
    except ValueError:
        return np.nan


def year_hits(row):
    hits = 0
    for v in row:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            continue
        if isinstance(v, (int, np.integer)) and 1900 <= int(v) <= 2100:
            hits += 1
        else:
            s = str(v).strip()
            if YEAR_RE.match(s):
                hits += 1
    return hits


def month_hits(row):
    hits = 0
    for v in row:
        if not isinstance(v, str):
            continue
        t = v.strip().lower()
        if not t:
            continue
        t = t.split()[0]  # czasem w komórce jest więcej tekstu
        if t in MONTH_MAP:
            hits += 1
    return hits


def pick_best_sheet(path: Path) -> str:
    """Wybiera arkusz, który ma najwięcej lat (2020.. itp.) w pierwszych ~80 wierszach."""
    xls = pd.ExcelFile(path)
    best_name = None
    best_score = -1
    for name in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=name, header=None)
        score = 0
        for r in range(min(80, len(df))):
            score = max(score, year_hits(df.iloc[r].tolist()))
        if score > best_score:
            best_score = score
            best_name = name
    if best_name is None:
        raise ValueError("Nie udało się wybrać arkusza z danymi.")
    return best_name


def find_year_row(df: pd.DataFrame) -> int:
    best_r, best_h = None, -1
    for r in range(min(120, len(df))):
        h = year_hits(df.iloc[r].tolist())
        if h > best_h:
            best_h, best_r = h, r
    if best_r is None or best_h < 3:
        raise ValueError("Nie znalazłem wiersza z latami (np. 2020..2024).")
    return best_r


def find_month_row(df: pd.DataFrame, year_row: int) -> int:
    best_r, best_h = 0, -1
    for r in range(0, max(1, year_row)):
        h = month_hits(df.iloc[r].tolist())
        if h > best_h:
            best_h, best_r = h, r
    return best_r


def find_data_col_start(df: pd.DataFrame, year_row: int) -> int:
    row = df.iloc[year_row, :].tolist()
    for c, v in enumerate(row):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            continue
        if isinstance(v, (int, np.integer)) and 1900 <= int(v) <= 2100:
            return c
        s = str(v).strip()
        if YEAR_RE.match(s):
            return c
    raise ValueError("Nie znalazłem pierwszej kolumny danych (pierwszego roku w wierszu lat).")


def find_suma_row(df: pd.DataFrame) -> int:
    """
    Szuka wiersza, gdzie w kolumnie 'Nazwa' (zwykle kol. 1) jest SUMA.
    Jeśli nie znajdzie w kolumnie 1, próbuje w kolumnie 0.
    """
    # Spróbuj najpierw kol. 1 (Kod, Nazwa)
    for col in [1, 0]:
        if col >= df.shape[1]:
            continue
        col_vals = df.iloc[:, col].astype(str).str.strip().str.lower()
        # dokładnie "suma" lub komórka zawierająca "suma"
        matches = col_vals.eq("suma") | col_vals.str.fullmatch(r"suma", na=False) | col_vals.str.contains(r"\bsuma\b", regex=True, na=False)
        idx = np.where(matches.to_numpy())[0]
        if len(idx) > 0:
            return int(idx[0])
    raise ValueError("Nie znalazłem wiersza 'SUMA'. Jeśli go nie ma, trzeba liczyć sumy z województw.")


def build_period_labels(months_row, years_row):
    # forward-fill miesiąca
    months = []
    current = None
    for m in months_row:
        if isinstance(m, str) and m.strip():
            t = m.strip().lower().split()[0]
            if t in MONTH_MAP:
                current = t
        months.append(current)

    years = []
    for y in years_row:
        if y is None or (isinstance(y, float) and np.isnan(y)):
            years.append(None)
        elif isinstance(y, (int, np.integer)):
            years.append(int(y))
        else:
            s = str(y).strip()
            years.append(int(s) if YEAR_RE.match(s) else None)

    periods = []
    for m, y in zip(months, years):
        if m is None or y is None:
            periods.append(None)
        else:
            periods.append((int(y), MONTH_MAP[m]))
    return periods


def main():
    in_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_INPUT)
    if not in_path.exists():
        raise FileNotFoundError(f"Nie widzę pliku: {in_path.resolve()}")

    sheet_in = pick_best_sheet(in_path)
    df = pd.read_excel(in_path, sheet_name=sheet_in, header=None)

    year_row = find_year_row(df)
    month_row = find_month_row(df, year_row)
    data_col_start = find_data_col_start(df, year_row)

    # okresy dla kolumn
    months_row = df.iloc[month_row, data_col_start:].tolist()
    years_row  = df.iloc[year_row,  data_col_start:].tolist()
    periods = build_period_labels(months_row, years_row)

    # wiersz SUMA i jego wartości
    suma_row = find_suma_row(df)
    suma_values_raw = df.iloc[suma_row, data_col_start:].tolist()
    suma_values = [to_number_pl(v) for v in suma_values_raw]

    out = []
    for p, v in zip(periods, suma_values):
        if p is None:
            continue
        y, m = p
        if OUTPUT_PERIOD_FORMAT == "MM.YYYY":
            label = f"{m:02d}.{y:04d}"
        else:
            label = f"{y:04d}-{m:02d}"
        out.append((y, m, label, None if np.isnan(v) else float(v)))

    # sort chronologiczny
    out.sort(key=lambda t: (t[0], t[1]))

    # zapis: kopia pliku + nowy arkusz
    out_path = in_path.with_name(in_path.stem + OUTPUT_SUFFIX + in_path.suffix)
    shutil.copy(in_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    if SHEET_OUT in wb.sheetnames:
        del wb[SHEET_OUT]
    ws = wb.create_sheet(SHEET_OUT)

    ws.append(["okres", "suma"])
    for _, __, label, val in out:
        ws.append([label, val])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18

    wb.save(out_path)
    print(f"OK: {out_path.name}")
    print(f"Wejściowy arkusz: {sheet_in}")
    print(f"Wiersz SUMA: {suma_row + 1} (Excel liczy od 1)")
    print(f"Arkusz wynikowy: {SHEET_OUT}")


if __name__ == "__main__":
    main()
